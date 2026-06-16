#!/usr/bin/env python3
"""
Send Overdue Pickup Notification Emails
Notifies about orders that are in Accepted status and past their scheduled pickup date
by more than 2 business days.
"""

import csv
import os
from typing import List, Dict, Optional, Set
from datetime import datetime, timedelta
from outlook_email_sender import OutlookEmailSender
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class OverduePickupNotifier:
    DISPATCHER_EMAILS = {
        'Kyle Bennett': 'kyle.bennett@hertz.com',
        'Matthew Coloian': 'matthew.coloian@hertz.com',
        'David Lagemann': 'david.lagemann@hertz.com',
        'Jason Payne': 'jason.payne@hertz.com',
        'Joe Tingelhoff': 'joe.tingelhoff@hertz.com',
        'Caleb Yates': 'caleb.yates@hertz.com',
        'Kelsie Zehender': 'kelsie.zehender@hertz.com'
    }

    TEAM_EMAILS = [
        'kyle.bennett@hertz.com',
        'matthew.coloian@hertz.com',
        'david.lagemann@hertz.com',
        'jason.payne@hertz.com',
        'joe.tingelhoff@hertz.com',
        'caleb.yates@hertz.com',
        'kelsie.zehender@hertz.com',
        'james.pham@hertz.com',
        'Ryan.Cappelen@hertz.com'
    ]

    def __init__(self, csv_file_path: str):
        self.csv_file_path = csv_file_path
        self.email_sender = OutlookEmailSender()
        self.overdue_orders = []

    def parse_date(self, date_string: str) -> Optional[datetime]:
        if not date_string or date_string.strip() == '':
            return None

        try:
            return datetime.strptime(date_string.strip(), '%m/%d/%Y')
        except ValueError:
            try:
                return datetime.strptime(date_string.strip(), '%Y-%m-%d')
            except ValueError:
                logger.warning(f"Could not parse date: {date_string}")
                return None

    @staticmethod
    def count_business_days(start_date: datetime, end_date: datetime) -> int:
        """Return the number of business days (Mon-Fri) between two dates, exclusive of start."""
        if start_date.date() >= end_date.date():
            return 0
        count = 0
        current = start_date + timedelta(days=1)
        while current.date() <= end_date.date():
            if current.weekday() < 5:  # Mon=0 .. Fri=4
                count += 1
            current += timedelta(days=1)
        return count

    def _is_impound_fee_only(self, vehicle_info: str) -> bool:
        """Return True if the vehicle description is just an impound fee with no real make/model."""
        if not vehicle_info:
            return True
        normalized = vehicle_info.strip().lower()
        if not normalized:
            return True
        impound_keywords = ['impound fee', 'impound fees', "impound fee's", 'storage fee', 'fees only']
        for kw in impound_keywords:
            if kw in normalized:
                return True
        return False

    def load_and_filter_orders(self) -> List[Dict]:
        """
        Load CSV data and filter for orders in Accepted status that are past
        their scheduled pickup date by more than 2 business days.
        Deduplicates so only one email is sent per VIN.
        Skips VINs whose vehicle description is just impound fees.
        """
        try:
            with open(self.csv_file_path, 'r', encoding='utf-8') as file:
                csv_reader = csv.DictReader(file)
                all_orders = list(csv_reader)

            logger.info(f"Loaded {len(all_orders)} total rows from CSV")

            today = datetime.now()
            seen_vins: Set[str] = set()
            filtered_orders = []

            for order in all_orders:
                vehicle_status = order.get('Vehicle Status', '').strip()
                carrier_name = order.get('Carrier Name', '').strip()
                pickup_date_str = order.get('Accepted by Carrier Date', '').strip()
                vin = order.get('VIN #', '').strip()
                vehicle_info = order.get('Vehicle Info', '').strip()

                if vehicle_status != 'Accepted':
                    continue

                if carrier_name.lower() == 'oem warranty tows':
                    continue

                if self._is_impound_fee_only(vehicle_info):
                    logger.info(f"Skipping VIN {vin} - vehicle description is impound fee only")
                    continue

                pickup_date = self.parse_date(pickup_date_str)
                if pickup_date is None:
                    continue

                biz_days_since_pickup = self.count_business_days(pickup_date, today)
                if biz_days_since_pickup < 2:
                    continue

                if vin and vin in seen_vins:
                    logger.info(f"Skipping duplicate VIN {vin}")
                    continue
                if vin:
                    seen_vins.add(vin)

                order['_biz_days_overdue'] = biz_days_since_pickup
                filtered_orders.append(order)

            self.overdue_orders = filtered_orders
            logger.info(f"Found {len(self.overdue_orders)} overdue orders (Accepted > 2 biz days past pickup, deduplicated by VIN)")
            return self.overdue_orders

        except Exception as e:
            logger.error(f"Failed to load and filter orders: {e}")
            raise

    def get_dispatcher_email(self, dispatcher_name: str) -> Optional[str]:
        return self.DISPATCHER_EMAILS.get(dispatcher_name.strip())

    def create_email_subject(self, order_data: Dict) -> str:
        order_id = order_data.get('Order ID', 'Unknown')
        return f"PAST PICKUP DATE ALERT - Order #{order_id}"

    def create_email_body(self, order_data: Dict) -> str:
        order_id = order_data.get('Order ID', 'N/A')
        vin = order_data.get('VIN #', 'N/A')
        vehicle_info = order_data.get('Vehicle Info', 'N/A')
        vehicle_status = order_data.get('Vehicle Status', 'N/A')

        pickup_business = order_data.get('Pickup Business Name', 'N/A')
        pickup_address = order_data.get('Pickup Address', 'N/A')
        pickup_city = order_data.get('Pickup City', 'N/A')
        pickup_state = order_data.get('Pickup State', 'N/A')
        pickup_zip = order_data.get('Pickup ZIP', 'N/A')

        delivery_business = order_data.get('Delivery Business Name', 'N/A')
        delivery_address = order_data.get('Delivery Address', 'N/A')
        delivery_city = order_data.get('Delivery City', 'N/A')
        delivery_state = order_data.get('Delivery State', 'N/A')
        delivery_zip = order_data.get('Delivery ZIP', 'N/A')

        scheduled_pickup = order_data.get('Accepted by Carrier Date', 'N/A')

        biz_days_overdue = order_data.get('_biz_days_overdue', 0)
        if biz_days_overdue > 0:
            overdue_text = f" ({biz_days_overdue} business day{'s' if biz_days_overdue != 1 else ''} overdue)"
        else:
            overdue_text = ""

        carrier_name = order_data.get('Carrier Name', 'N/A')

        email_body = f"""Hello {carrier_name},

This order is still showing in Accepted status on Super Dispatch.

Please provide the ETA for pickup. Also, let us know if there are any issues that may be affecting pickup or delivery.

Kindly REPLY ALL to this email with a status update when possible.

ORDER DETAILS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Order ID: {order_id}
Current Status: {vehicle_status}
Scheduled Pickup: {scheduled_pickup}{overdue_text}
VIN #: {vin}
Vehicle: {vehicle_info}

PICKUP LOCATION:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Business Name: {pickup_business}
Address: {pickup_address}, {pickup_city}, {pickup_state} {pickup_zip}

DELIVERY LOCATION:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Business Name: {delivery_business}
Address: {delivery_address}, {delivery_city}, {delivery_state} {delivery_zip}

ACTION REQUIRED:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
This order is past the scheduled pickup date. Please respond with one of the following:

1. Your updated pickup ETA
2. Current status and any issues preventing pickup
3. If you need this order to be reassigned to another carrier

We understand that delays can occur, and we appreciate your prompt communication to help us keep our customers informed.



Thank you for your continued partnership.

Best regards,
Hertz Impound Dispatch Team
AssetRecovery@hertz.com"""

        return email_body

    def preview_emails(self) -> None:
        if not self.overdue_orders:
            print("No overdue orders found.")
            return

        print("\n" + "="*80)
        print(f"PREVIEW: {len(self.overdue_orders)} PAST PICKUP DATE NOTIFICATIONS")
        print("="*80)

        for i, order in enumerate(self.overdue_orders, 1):
            order_id = order.get('Order ID', 'N/A')
            carrier_name = order.get('Carrier Name', 'N/A')
            carrier_email = order.get('Carrier Email', 'N/A')
            dispatcher = order.get('Dispatcher', 'N/A')
            dispatcher_email = self.get_dispatcher_email(dispatcher)
            pickup_date = order.get('Accepted by Carrier Date', 'N/A')
            vehicle_status = order.get('Vehicle Status', 'N/A')
            biz_days = order.get('_biz_days_overdue', 0)

            print(f"\n{i}. Order #{order_id}")
            print(f"   Status: {vehicle_status}")
            print(f"   Carrier: {carrier_name} ({carrier_email})")
            print(f"   Scheduled Pickup: {pickup_date} ({biz_days} business days overdue)")
            print(f"   Dispatcher: {dispatcher}")
            if dispatcher_email:
                print(f"   Will CC: {dispatcher_email}")
            else:
                print(f"   [WARNING] No email found for dispatcher: {dispatcher}")

        if self.overdue_orders:
            print("\n" + "="*80)
            print("SAMPLE EMAIL (First Order)")
            print("="*80)
            sample_order = self.overdue_orders[0]
            dispatcher = sample_order.get('Dispatcher', '')
            dispatcher_email = self.get_dispatcher_email(dispatcher)

            print(f"FROM: AssetRecovery@hertz.com")
            print(f"TO: {sample_order.get('Carrier Email', 'N/A')}")
            if dispatcher_email:
                print(f"CC: {dispatcher_email} ({dispatcher})")
            print(f"SUBJECT: {self.create_email_subject(sample_order)}")
            print("\nBODY:")
            print("-" * 80)
            print(self.create_email_body(sample_order))
            print("="*80)

    def send_emails(self, preview_mode: bool = True) -> Dict:
        if not self.overdue_orders:
            return {
                'total_orders': 0,
                'sent_successfully': 0,
                'failed_to_send': 0,
                'skipped': 0,
                'details': []
            }

        results = {
            'total_orders': len(self.overdue_orders),
            'sent_successfully': 0,
            'failed_to_send': 0,
            'skipped': 0,
            'details': []
        }

        for order in self.overdue_orders:
            order_id = order.get('Order ID', 'Unknown')
            carrier_email = order.get('Carrier Email', '').strip()
            carrier_name = order.get('Carrier Name', 'Unknown Carrier')
            dispatcher = order.get('Dispatcher', '').strip()
            dispatcher_email = self.get_dispatcher_email(dispatcher)

            if not carrier_email or carrier_email.lower() in ['n/a', 'na', '']:
                results['skipped'] += 1
                results['details'].append({
                    'order_id': order_id,
                    'carrier_name': carrier_name,
                    'status': 'SKIPPED - No carrier email',
                    'email': carrier_email,
                    'dispatcher': dispatcher,
                    'cc': None
                })
                continue

            subject = self.create_email_subject(order)
            body = self.create_email_body(order)

            if preview_mode:
                results['details'].append({
                    'order_id': order_id,
                    'carrier_name': carrier_name,
                    'status': 'PREVIEW - Would send',
                    'email': carrier_email,
                    'dispatcher': dispatcher,
                    'cc': dispatcher_email
                })
                continue

            cc_recipients = []
            if dispatcher_email:
                cc_recipients.append(dispatcher_email)

            try:
                success = self.email_sender.send_email(
                    to_recipients=[carrier_email],
                    subject=subject,
                    body=body,
                    cc_recipients=cc_recipients if cc_recipients else None,
                    send_from='AssetRecovery@hertz.com'
                )

                if success:
                    results['sent_successfully'] += 1
                    status = 'SENT SUCCESSFULLY'
                else:
                    results['failed_to_send'] += 1
                    status = 'FAILED TO SEND'

                results['details'].append({
                    'order_id': order_id,
                    'carrier_name': carrier_name,
                    'status': status,
                    'email': carrier_email,
                    'dispatcher': dispatcher,
                    'cc': dispatcher_email
                })

            except Exception as e:
                results['failed_to_send'] += 1
                results['details'].append({
                    'order_id': order_id,
                    'carrier_name': carrier_name,
                    'status': f'ERROR: {str(e)}',
                    'email': carrier_email,
                    'dispatcher': dispatcher,
                    'cc': dispatcher_email
                })
                logger.error(f"Failed to send email for order {order_id}: {e}")

        return results

    def export_to_excel(self, results: Dict, filename: str = None) -> str:
        try:
            if filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'overdue_pickup_emails_{timestamp}.xlsx'

            wb = Workbook()
            ws = wb.active
            ws.title = "Overdue Pickup Orders"

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            headers = [
                'Dispatcher', 'Order ID', 'Status', 'Scheduled Pickup',
                'Biz Days Overdue', 'VIN', 'Vehicle Info', 'Carrier Name',
                'Carrier Email', 'Pickup Name', 'Pickup Address',
                'Delivery Name', 'Delivery Address', 'Email Status', 'Remarks'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            for idx, detail in enumerate(results['details'], 2):
                order_data = None
                for order in self.overdue_orders:
                    if order.get('Order ID') == detail['order_id']:
                        order_data = order
                        break

                biz_days_overdue = order_data.get('_biz_days_overdue', '') if order_data else ''

                status = detail['status']
                if 'SUCCESSFULLY' in status:
                    status_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_display = "Sent"
                elif 'FAILED' in status or 'ERROR' in status:
                    status_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_display = "Failed"
                else:
                    status_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    status_display = "Skipped"

                scheduled_pickup = order_data.get('Accepted by Carrier Date', 'N/A') if order_data else 'N/A'
                vin = order_data.get('VIN #', 'N/A') if order_data else 'N/A'
                vehicle_info = order_data.get('Vehicle Info', 'N/A') if order_data else 'N/A'

                pickup_business = order_data.get('Pickup Business Name', 'N/A') if order_data else 'N/A'
                pickup_full_address = ''
                if order_data:
                    pickup_addr = order_data.get('Pickup Address', '')
                    pickup_city = order_data.get('Pickup City', '')
                    pickup_state = order_data.get('Pickup State', '')
                    pickup_zip = order_data.get('Pickup ZIP', '')
                    pickup_full_address = f"{pickup_addr}, {pickup_city}, {pickup_state} {pickup_zip}" if pickup_addr else 'N/A'

                delivery_business = order_data.get('Delivery Business Name', 'N/A') if order_data else 'N/A'
                delivery_full_address = ''
                if order_data:
                    delivery_addr = order_data.get('Delivery Address', '')
                    delivery_city = order_data.get('Delivery City', '')
                    delivery_state = order_data.get('Delivery State', '')
                    delivery_zip = order_data.get('Delivery ZIP', '')
                    delivery_full_address = f"{delivery_addr}, {delivery_city}, {delivery_state} {delivery_zip}" if delivery_addr else 'N/A'

                row_data = [
                    detail['dispatcher'],
                    detail['order_id'],
                    order_data.get('Vehicle Status', 'N/A') if order_data else 'N/A',
                    scheduled_pickup,
                    biz_days_overdue,
                    vin,
                    vehicle_info,
                    detail['carrier_name'],
                    detail['email'],
                    pickup_business,
                    pickup_full_address,
                    delivery_business,
                    delivery_full_address,
                    status_display,
                    status if status_display != "Sent" else ''
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=idx, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    if col == 14:
                        cell.fill = status_fill
                        cell.font = Font(bold=True)

            column_widths = [22, 15, 12, 22, 18, 22, 30, 38, 35, 35, 50, 40, 55, 15, 35]
            for i, width in enumerate(column_widths, 1):
                col_letter = chr(64 + i) if i <= 26 else chr(64 + (i - 1) // 26) + chr(65 + (i - 1) % 26)
                ws.column_dimensions[col_letter].width = width

            ws.auto_filter.ref = f"A1:O{len(results['details']) + 1}"

            summary_row = len(results['details']) + 3
            ws.cell(row=summary_row, column=1, value="SUMMARY:").font = Font(bold=True, size=12)

            ws.cell(row=summary_row + 1, column=1, value="Total Orders:")
            ws.cell(row=summary_row + 1, column=2, value=results['total_orders'])

            ws.cell(row=summary_row + 2, column=1, value="Sent Successfully:")
            ws.cell(row=summary_row + 2, column=2, value=results['sent_successfully'])
            ws.cell(row=summary_row + 2, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            ws.cell(row=summary_row + 3, column=1, value="Failed:")
            ws.cell(row=summary_row + 3, column=2, value=results['failed_to_send'])
            if results['failed_to_send'] > 0:
                ws.cell(row=summary_row + 3, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            ws.cell(row=summary_row + 4, column=1, value="Skipped:")
            ws.cell(row=summary_row + 4, column=2, value=results['skipped'])
            if results['skipped'] > 0:
                ws.cell(row=summary_row + 4, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            wb.save(filename)
            logger.info(f"Excel report saved to: {filename}")

            return filename

        except Exception as e:
            logger.error(f"Failed to create Excel report: {e}")
            return None

    def send_summary_email(self, excel_file: str, results: Dict, send_email: bool = True) -> bool:
        try:
            timestamp = datetime.now().strftime('%B %d, %Y at %I:%M %p')
            subject = f"Past Pickup Date Notifications Sent - {datetime.now().strftime('%m/%d/%Y')}"

            dispatcher_counts = {}
            for detail in results['details']:
                dp = detail.get('dispatcher', 'Unassigned')
                if not dp or dp.strip() == '' or dp == 'N/A':
                    dp = 'Unassigned'
                dispatcher_counts[dp] = dispatcher_counts.get(dp, 0) + 1

            body = f"""Team,

Past pickup date notification emails have been processed on {timestamp}.

SUMMARY:
Total Orders Processed: {results['total_orders']}
Emails Sent Successfully: {results['sent_successfully']}
Failed to Send: {results['failed_to_send']}
Skipped: {results['skipped']}

ORDERS BY DISPATCHER:
"""

            for dp in sorted(dispatcher_counts.keys()):
                count = dispatcher_counts[dp]
                body += f"  {dp}: {count}\n"

            body += f"""
ATTACHED:
The attached Excel file contains complete details of all overdue pickup orders processed.

Please review the attached report for full details on your assigned orders.

Best regards,
Hertz Impound Dispatch Team
AssetRecovery@hertz.com"""

            if send_email:
                full_path = os.path.abspath(excel_file)
                success = self.email_sender.send_email(
                    to_recipients=self.TEAM_EMAILS,
                    subject=subject,
                    body=body,
                    attachments=[full_path],
                    send_from='AssetRecovery@hertz.com'
                )

                if success:
                    logger.info(f"Summary email sent to {len(self.TEAM_EMAILS)} dispatchers")
                    return True
                else:
                    logger.error("Failed to send summary email")
                    return False
            else:
                print("\n" + "="*80)
                print("PREVIEW: SUMMARY EMAIL TO DISPATCHERS")
                print("="*80)
                print(f"FROM: AssetRecovery@hertz.com")
                print(f"TO: {', '.join(self.TEAM_EMAILS[:3])}...")
                print(f"    (Total: {len(self.TEAM_EMAILS)} dispatchers)")
                print(f"SUBJECT: {subject}")
                print(f"ATTACHMENT: {excel_file}")
                print("\nBODY:")
                print("-" * 80)
                print(body)
                print("="*80)
                print("\n[INFO] Email NOT sent (preview mode)")
                return True

        except Exception as e:
            logger.error(f"Failed to send summary email: {e}")
            return False


def main():
    """Main function to send overdue pickup notifications."""

    csv_file = 'raw.csv'

    if not os.path.exists(csv_file):
        print(f"[ERROR] CSV file '{csv_file}' not found in current directory")
        return

    print(f"[INFO] Using CSV file: {csv_file}")

    try:
        notifier = OverduePickupNotifier(csv_file)

        print("[INFO] Scanning for orders past pickup date (Accepted > 2 business days)...")
        notifier.load_and_filter_orders()

        if not notifier.overdue_orders:
            print("[SUCCESS] No overdue pickups found! All accepted orders are within the 2 business day window.")
            return

        print(f"\n[WARNING] Found {len(notifier.overdue_orders)} orders past pickup date")

        notifier.preview_emails()

        print("\n" + "="*80)
        print("[WARNING] READY TO SEND PAST PICKUP DATE NOTIFICATIONS")
        print("="*80)
        print(f"This will send {len(notifier.overdue_orders)} emails to carriers")
        print("FROM: AssetRecovery@hertz.com")
        print("CC: Respective dispatcher for each order")
        print("\n[IMPORTANT] Each email will be sent to the carrier with the dispatcher CC'd.")

        confirm = input("\nType 'SEND' to confirm and send all emails (or anything else to cancel): ").strip()

        if confirm.upper() == 'SEND':
            print("\n[INFO] Sending past pickup date notifications...")

            results = notifier.send_emails(preview_mode=False)

            print("\n" + "="*80)
            print("[RESULTS] EMAIL SENDING RESULTS")
            print("="*80)
            print(f"Total Orders: {results['total_orders']}")
            print(f"[SUCCESS] Sent Successfully: {results['sent_successfully']}")
            print(f"[FAILED] Failed to Send: {results['failed_to_send']}")
            print(f"[SKIPPED] Skipped: {results['skipped']}")

            if results['details']:
                print("\nDETAILS:")
                for detail in results['details']:
                    status_prefix = "[SUCCESS]" if "SUCCESSFULLY" in detail['status'] else "[FAILED]" if "FAILED" in detail['status'] or "ERROR" in detail['status'] else "[SKIPPED]"
                    cc_info = f" | CC: {detail['cc']}" if detail['cc'] else ""
                    print(f"  {status_prefix} Order #{detail['order_id']} - {detail['carrier_name']} ({detail['email']}){cc_info}")

            if results['sent_successfully'] > 0:
                print(f"\n[SUCCESS] Successfully sent {results['sent_successfully']} past pickup date notifications!")

            if results['failed_to_send'] > 0:
                print(f"\n[WARNING] {results['failed_to_send']} emails failed to send. Check the details above.")

            print("\n" + "="*80)
            print("[INFO] Creating Excel report...")
            excel_file = notifier.export_to_excel(results)
            if excel_file:
                full_path = os.path.abspath(excel_file)
                print(f"[SUCCESS] Excel report created: {excel_file}")
                print(f"[INFO] Full path: {full_path}")

                print("\n[INFO] Sending summary email to dispatch team...")
                summary_sent = notifier.send_summary_email(excel_file, results, send_email=True)
                if summary_sent:
                    print("[SUCCESS] Summary email with Excel report sent to dispatch team")
                else:
                    print("[WARNING] Failed to send summary email to dispatch team")
            else:
                print("[WARNING] Failed to create Excel report")
        else:
            print("[CANCELLED] Email sending cancelled. No emails were sent.")

    except Exception as e:
        print(f"[ERROR] Error: {e}")
        logger.error(f"Application error: {e}", exc_info=True)


if __name__ == "__main__":
    main()
